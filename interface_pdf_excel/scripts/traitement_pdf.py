

import pandas as pd
import os
import cv2
import numpy as np
import pytesseract
import re
from pdf2image import convert_from_path
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
from tkinter import filedialog, Tk
import time
start_time = time.time()

pytesseract.pytesseract_cmd = "/usr/bin/tesseract"

# --- OCR utils ---
def extract_block_text(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)
    _, threshed = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    config = r'--oem 3 --psm 6 -l fra+eng'
    return pytesseract.image_to_string(threshed, config=config).strip()

# --- Détection des titres et sous-titres ---

def remove_duplicate_realisations(text):
    if not isinstance(text, str):
        return text

    text = text.replace("\xa0", " ")  # espace insécable
    text = re.sub(r"\s+", " ", text).strip()

    # Toutes les occurrences
    matches = re.findall(
        r"(Nombre\s+de\s+r[ée]alisation\s*:?\s*\d+\s*/\s*\d+)",
        text, flags=re.IGNORECASE
    )

    if len(matches) > 1:
        # Supprimer toutes
        text = re.sub(
            r"(Nombre\s+de\s+r[ée]alisation\s*:?\s*\d+\s*/\s*\d+)",
            "", text, flags=re.IGNORECASE
        ).strip()
        # Réinsertion d’une seule
        text = f"{text} {matches[0]}".strip()

    return text






def detect_pdfplumber_essai_titles_with_y(pdf_page, page_number, img_height, pdf_path):
    from collections import defaultdict
    import re
    from pdf2image import convert_from_path

    essai_titles = []
    seen_titles = set()

    words = pdf_page.extract_words()
    if not words:
        return essai_titles

    lines_by_y = defaultdict(list)
    for word in words:
        y = round(word["top"])
        lines_by_y[y].append((word["x0"], word["text"]))

    sorted_y = sorted(lines_by_y.keys())

    for idx, y in enumerate(sorted_y):
        line_text = " ".join(txt for _, txt in sorted(lines_by_y[y])).strip()
        line_text = line_text.replace("Essain", "Essai")
        line_text = remove_duplicate_realisations(line_text)

        match_num = re.search(r"Essai\s*n[\u00ba\u00b0]?\s*(\d+)", line_text, re.IGNORECASE)
        if not match_num:
            continue

        num = match_num.group(1).strip()
        nb_rea = ""

        # Recherche dans la même ligne
        match_nb_rea = re.findall(r"Nombre\s+de\s+r[ée]alisation\s*:?\s*(\d+\s*/\s*\d+)", line_text, re.IGNORECASE)
        if match_nb_rea:
            nb_rea = match_nb_rea[0].strip()

        # Si pas trouvé, chercher dans les lignes suivantes
        if not nb_rea:
            for dy in sorted_y[idx + 1:idx + 6]:
                candidate = " ".join(txt for _, txt in sorted(lines_by_y[dy])).strip()
                candidate = remove_duplicate_realisations(candidate)
                match_follow = re.findall(r"Nombre\s+de\s+r[ée]alisation\s*:?\s*(\d+\s*/\s*\d+)", candidate, re.IGNORECASE)
                if match_follow:
                    nb_rea = match_follow[0].strip()
                    break

        # Nettoyage nom
        nom = re.sub(r"Essai\s*n[\u00ba\u00b0]?\s*" + re.escape(num), "", line_text, flags=re.IGNORECASE)
        nom = re.sub(r"(Nombre\s+de\s+r[ée]alisation\s*:?\s*\d+\s*/\s*\d+)", "", nom, flags=re.IGNORECASE)
        nom = nom.strip()

        full_line = f"Essai n°{num}: {nom}"
        if nb_rea:
            full_line += f" Nombre de réalisation : {nb_rea}"

        # 🧼 NC:\> pip install pandasettoyage final
        full_line = remove_duplicate_realisations(full_line)

        key = f"{num}-{nb_rea or nom}"
        if key in seen_titles:
            continue
        seen_titles.add(key)

        y_img = int((y / pdf_page.height) * img_height)
        essai_titles.append((y_img, 'title', full_line, num, nom, nb_rea))

    return essai_titles























def detect_titles_with_y(img, pdf_page, page_number, img_height, pdf_path):
    import cv2
    import numpy as np
    import pytesseract
    import re

    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    height = img.shape[0]
    titles = []

    def extract_block_text(local_img):
        gray = cv2.cvtColor(local_img, cv2.COLOR_BGR2GRAY)
        gray = cv2.bilateralFilter(gray, 9, 75, 75)
        _, threshed = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        config = r'--oem 3 --psm 4 -l fra+eng'
        return pytesseract.image_to_string(threshed, config=config).strip()

    # Détection ANNEXES
    for y in range(0, height - 100, 50):
        segment = img[y:y+100, :]
        text = extract_block_text(segment).lower()
        if "annexes" in text:
            titles.append((y, 'title', "ANNEXES"))
            break

    # Détection blocs violets
    lower_violet = np.array([120, 40, 40])
    upper_violet = np.array([155, 255, 255])
    mask = cv2.inRange(hsv, lower_violet, upper_violet)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((5, 5), np.uint8))
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        if w > 60 and h > 20:
            roi = img[y:y+h, x:x+w]
            roi_gray = cv2.bitwise_not(cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY))
            text = pytesseract.image_to_string(roi_gray, lang='fra+eng', config='--psm 6').strip()
            if text:
                titles.append((y, "title", text))

    # Détection sous-titres via pdfplumber
    for full_entry in detect_pdfplumber_essai_titles_with_y(pdf_page, page_number, img_height, pdf_path):
        y_img, typ, full_line, num, nom, nb_rea = full_entry
        clean_line = remove_duplicate_realisations(full_line)
        titles.append((y_img, typ, clean_line))  # ✅ Nettoyage appliqué ici

    return sorted(titles, key=lambda x: x[0]), height



def merge_tables_between_titles(blocks):
    merged_blocks = []
    last_title = None
    current_table_rows = []

    for block in blocks:
        if block[1] == 'title':
            # Nouveau titre détecté
            if current_table_rows and last_title:
                merged_blocks.append((last_title[0], 'title', last_title[2]))
                merged_blocks.append((last_title[0] + 0.1, 'table', current_table_rows))
                current_table_rows = []

            # 🔁 Important : on traite chaque titre comme unique
            last_title = block

        elif block[1] == 'table':
            if block[2] and isinstance(block[2], list):
                current_table_rows.extend(block[2])

    # Dernier bloc
    if current_table_rows and last_title:
        merged_blocks.append((last_title[0], 'title', last_title[2]))
        merged_blocks.append((last_title[0] + 0.1, 'table', current_table_rows))

    return merged_blocks



def create_fused_results_sheet(extracted_excel_path, log_func=print):
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter

    log_func("\n📌 Fusion des tableaux entre chaque paire de titres...")

    wb = load_workbook(extracted_excel_path)
    ws_in = wb["Résultat Ordonné"]
    ws_out = wb.create_sheet("Résultat Fusionné")

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill("solid", fgColor="FFFACD")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font_style = Font(name='Calibri', size=11)

    # Étape 1 : lire tous les blocs (titre ou tableau)
    blocks = []
    y = 0
    row = 1
    while row <= ws_in.max_row:
        cell = ws_in.cell(row=row, column=1)
        val = str(cell.value).strip() if cell.value else ""
        val_lower = val.lower()

        # Titre ?
        is_title = False
        if "essai n" in val_lower:
            is_title = True
        elif "nombre de réalisation" in val_lower:
            is_title = True
        elif "annexes" in val_lower:
            is_title = True
        elif len(val) > 0 and ws_in.cell(row=row, column=2).value is None:
            is_title = True

        if is_title:
            blocks.append((y, 'title', val))
            row += 1

        # Tableau avec entête
        elif val_lower in ["nom de la mesure", "nom"] and str(ws_in.cell(row=row, column=2).value).lower() in ["valeur de la mesure", "valeur"]:
            header = [ws_in.cell(row=row, column=1).value, ws_in.cell(row=row, column=2).value]
            table = []
            row += 1
            while row <= ws_in.max_row:
                v1 = ws_in.cell(row=row, column=1).value
                v2 = ws_in.cell(row=row, column=2).value
                if not v1 and not v2:
                    break
                table.append([v1, v2])
                row += 1
            if table:
                table.insert(0, header)
                blocks.append((y + 0.1, 'table', table))

        else:
            table = []
            for _ in range(10):
                v1 = ws_in.cell(row=row, column=1).value
                v2 = ws_in.cell(row=row, column=2).value
                if not v1 and not v2:
                    break
                table.append([v1, v2])
                row += 1
            if table:
                blocks.append((y + 0.1, 'table', table))
            else:
                row += 1
        y += 1

    # Fusion des tableaux entre titres
    fused_blocks = merge_tables_between_titles(blocks)

    # Écriture finale
    row_offset = 1
    for block in fused_blocks:
        if block[1] == 'title':
            clean_val = remove_duplicate_realisations(block[2])  # ✅ nettoyage ici
            ws_out.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=5)
            cell = ws_out.cell(row=row_offset, column=1, value=clean_val)
            cell.font = Font(bold=True, italic=True)
            row_offset += 1
        elif block[1] == 'table':
            for r_idx, row in enumerate(block[2]):
                for c_idx, val in enumerate(row):
                    cell = ws_out.cell(row=row_offset + r_idx, column=c_idx + 1, value=val if val else "")
                    cell.border = border
                    cell.fill = fill
                    cell.alignment = align_center
                    cell.font = font_style
                    ws_out.column_dimensions[get_column_letter(c_idx + 1)].width = 25
            row_offset += len(block[2]) + 2

    wb.save(extracted_excel_path)
    log_func(f"\n✅ Feuille 'Résultat Fusionné' créée avec succès dans : {extracted_excel_path}")











def filter_duplicate_titles(titles, y_threshold=25):
    import difflib
    filtered = []
    seen_texts = []
    for y, typ, content in sorted(titles, key=lambda x: x[0]):
        content_clean = content.lower().strip()
        is_duplicate = False
        for prev_y, prev_text in seen_texts:
            if abs(y - prev_y) < y_threshold:
                similarity = difflib.SequenceMatcher(None, content_clean, prev_text).ratio()
                if similarity > 0.9:
                    is_duplicate = True
                    break
        if not is_duplicate:
            filtered.append((y, typ, content))
            seen_texts.append((y, content_clean))
    return filtered


def is_structured_table(table_data, y_position=None):
    import re

    if not table_data:
        return False

    # ⚠️ Option : ignorer les tableaux trop hauts dans la page (entêtes décoratifs)
    if y_position is not None and y_position < 50:
        return False

    total_cells = sum(len(row) for row in table_data)
    if total_cells == 0:
        return False

    # ⚠️ Exception pour les tableaux à une seule ligne contenant une mesure simple
    if len(table_data) == 1:
        first_row = table_data[0]
        if len(first_row) >= 2:
            nom, val = str(first_row[0]).lower(), str(first_row[1]).lower()
            if nom and val:
                # ✅ Autorise une ligne du type "Nom de la mesure", "Valeur"
                if re.search(r'(valeur|mesure|temp|pression|vitesse|niveau|tension|fréquence|temps)', nom):
                    if re.search(r'\d', val):  # contient au moins un chiffre
                        return True

    # Texte complet pour analyse heuristique
    full_text = " ".join(str(cell).lower() for row in table_data for cell in row if cell)

    # Cas décoratif à exclure (rame Zxxxxx, masj, etc.)
    if re.search(r"\bz\s*\d{4,6}\b", full_text) and "masj" in full_text:
        return False

    # Analyse simple du contenu
    useful_cells = sum(
        1 for row in table_data for cell in row
        if cell and re.search(r'\d{2,}|\d+\.\d+|bar|oui|non|°c|v|hz|mm|ms|kg|s', str(cell).lower())
    )

    if useful_cells / total_cells < 0.2:
        return False

    return True




# --- Traitement principal : extraction titres + tableaux fusionnés ---
def extract_and_order_blocks(pdf_path, output_excel, log_func=print):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
    from openpyxl.utils import get_column_letter

    pages_img = convert_from_path(pdf_path, dpi=300)
    pdf = pdfplumber.open(pdf_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Résultat Ordonné"

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))
    fill = PatternFill("solid", fgColor="FFFACD")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font_style = Font(name='Calibri', size=11)

    row_offset = 1

    for i, (img_pil, pdf_page) in enumerate(zip(pages_img, pdf.pages)):
        log_func(f"\n[INFO] Page {i+1}")
        img_cv = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
        pdf_height = pdf_page.height
        img_height = img_cv.shape[0]

        titles, _ = detect_titles_with_y(img_cv, pdf_page, i + 1, img_height, pdf_path)

        blocks = []

        # Convertir Y image → Y PDF
        for y_img, typ, content in titles:
            y_ratio = y_img / img_height
            y_pdf = y_ratio * pdf_height
            blocks.append((y_pdf, 'title', content))

        for table in pdf_page.find_tables():
            y_top = table.bbox[3]
            blocks.append((y_top, 'table', table.extract()))

        # Trier tous les blocs verticalement
        blocks = sorted(blocks, key=lambda x: x[0])

        for y, typ, content in blocks:
            if typ == 'title':
                # 🧼 Appliquer le nettoyage ici
                content = remove_duplicate_realisations(content)
                log_func(f"  ➤ Titre à y={int(y)} : {content}")
                ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=5)
                cell = ws.cell(row=row_offset, column=1, value=content)
                cell.font = Font(bold=True, italic=True)
                row_offset += 1

            elif typ == 'table':
                log_func(f"  ➤ Tableau à y={int(y)}")
                table_data = content

                if is_structured_table(table_data, y_position=y):
                    max_cols = max(len(row) for row in table_data)
                    for r_idx, row in enumerate(table_data):
                        for c_idx, val in enumerate(row):
                            cell = ws.cell(row=row_offset + r_idx, column=c_idx + 1, value=val.strip() if val else "")
                            cell.border = border
                            cell.fill = fill
                            cell.alignment = align_center
                            cell.font = font_style
                            ws.column_dimensions[get_column_letter(c_idx + 1)].width = 25
                    row_offset += len(table_data) + 2

    wb.save(output_excel)
    log_func(f"\n✅ Fichier Excel généré : {output_excel}")






def extract_essais_and_mesures_from_titles(ws):
    mesures_essais = []

    for row_idx in range(1, ws.max_row):
        cell_val = str(ws.cell(row=row_idx, column=1).value or "").strip()
        match = re.match(
            r"Essai\s*n[\u00ba\u00b0]?\s*(\d+)\s*:\s*(.*?)\s*(?:Nombre de r[ée]alisation\s*:\s*([0-9]+/[0-9]+))?$",
            cell_val,
            re.IGNORECASE
        )

        if match:
            num_essai = match.group(1).strip()
            nom_essai = match.group(2).strip()
            nb_rea = match.group(3).strip() if match.group(3) else ""

            r = row_idx + 1
            # Vérifie si on a un tableau avec "Nom de la mesure" en-tête
            if (
                ws.cell(row=r, column=1).value == "Nom de la mesure"
                and ws.cell(row=r, column=2).value == "Valeur de la mesure"
            ):
                r += 1  # sauter l'en-tête
            # Sinon, c’est peut-être directement les mesures

            while r <= ws.max_row:
                nom = ws.cell(row=r, column=1).value
                val = ws.cell(row=r, column=2).value
                if not nom and not val:
                    break

                # Ne pas ajouter une ligne vide ou une ligne d'en-tête
                if str(nom).strip().lower() in ["nom de la mesure", "valeur de la mesure"]:
                    r += 1
                    continue

                mesures_essais.append({
                    "Numéro d'essai": num_essai,
                    "Nom de l'essai": nom_essai,
                    "Nombre de réalisation": nb_rea,
                    "Nom de la mesure": nom,
                    "Valeur de la mesure": val
                })
                r += 1

    return mesures_essais



def create_global_summary(extracted_excel_path, log_func=print):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import re

    wb = load_workbook(extracted_excel_path)
    ws = wb["Résultat Fusionné"]

    program_name = editor = test_date = rame_number = poste_number = gvg_date = ""

    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ""
            if "nom du programme" in val:
                program_name = row[1].value
            elif "rédacteur" in val:
                editor = row[1].value
            elif "date d'essai" in val:
                test_date = row[1].value
            elif "n° de la rame" in val:
                rame_number = row[1].value
            elif "n° du poste" in val:
                poste_number = row[1].value
            elif "date de validité gvg" in val:
                gvg_date = row[1].value

    essais_mesures = extract_essais_and_mesures_from_titles(ws)

    ws_summary = wb.create_sheet("Résumé Global")
    header = [
        "Nom du programme", "Rédacteur", "Date d'essai", "Numéro rame",
        "Numéro du poste", "Date GVG du poste",
        "Numéro d'essai", "Nom de l'essai", "Nombre de réalisation",
        "Nom de la mesure", "Valeur de la mesure"
    ]
    ws_summary.append(header)

    for m in essais_mesures:
        ws_summary.append([
            program_name, editor, test_date, rame_number, poste_number, gvg_date,
            m["Numéro d'essai"], m["Nom de l'essai"], m["Nombre de réalisation"],
            m["Nom de la mesure"], m["Valeur de la mesure"]
        ])

    # Mise en forme
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill("solid", fgColor="FFFF99")

    for col in ws_summary.columns:
        col_letter = col[0].column_letter
        ws_summary.column_dimensions[col_letter].width = 25
        for cell in col:
            cell.alignment = center
            cell.border = border
            if cell.row == 1:
                cell.font = bold_font
                cell.fill = fill

    wb.save(extracted_excel_path)
    log_func(f"✅ Résumé Global finalisé avec colonnes séparées dans : {extracted_excel_path}")









from tkinter import filedialog, Tk
import time
def fix_duplicate_realisations_in_excel(file_path, log_func=print):
    from openpyxl import load_workbook
    import re

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        log_func(f"❌ Impossible d'ouvrir le fichier {file_path} : {e}")
        return

    if "Résultat Fusionné" not in wb.sheetnames:
        log_func(f"❌ Feuille 'Résultat Fusionné' manquante dans {file_path}")
        return

    ws = wb["Résultat Fusionné"]
    corrected = False

    for row in ws.iter_rows(min_row=1, max_col=1):  # colonne A
        cell = row[0]
        val = str(cell.value or "").strip()
        val = val.replace("\xa0", " ")

        val_final = remove_duplicate_realisations(val)

        if val_final != val:
            cell.value = val_final
            corrected = True
            log_func(f"🛠️ Corrigé ligne {cell.row} ➤ {val} → {val_final}")

    if corrected:
        wb.save(file_path)
        log_func(f"✅ Corrections sauvegardées dans : {file_path}")
    else:
        log_func(f"✅ Aucun doublon 'Nombre de réalisation' à corriger dans : {file_path}")


def check_duplicate_realisations_in_excel(file_path, log_func=print):
    from openpyxl import load_workbook
    import re

    log_func("\n🔎 Vérification des doublons 'Nombre de réalisation'...")

    wb = load_workbook(file_path)
    if "Résultat Fusionné" not in wb.sheetnames:
        log_func("❌ Feuille 'Résultat Fusionné' non trouvée.")
        return

    ws = wb["Résultat Fusionné"]
    found = False

    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        val = str(cell.value or "")
        matches = re.findall(
            r"(Nombre\s+de\s+r[ée]alisation\s*:?\s*\d+/\d+)",
            val,
            flags=re.IGNORECASE
        )
        if len(matches) > 1:
            log_func(f"⚠️ Doublon détecté ligne {cell.row} ➤ {val}")
            found = True

    if not found:
        log_func("✅ Aucun doublon trouvé dans les titres.")


import os
def process_single_pdf(pdf_path, output_dir, log_func=print):
    output_excel = os.path.join(output_dir, os.path.basename(pdf_path).replace(".pdf", "_fusion_finale_TOPDOWN_OK.xlsx"))
    log_func(f"\n📄 Traitement : {os.path.basename(pdf_path)}")
    try:
        extract_and_order_blocks(pdf_path, output_excel)
        create_fused_results_sheet(output_excel)
        create_global_summary(output_excel)
        fix_duplicate_realisations_in_excel(output_excel)
        check_duplicate_realisations_in_excel(output_excel)
    except Exception as e:
        log_func(f"❌ Erreur lors du traitement de {pdf_path} : {e}")

# Réécriture de la fonction avec gestion des logs à retourner à Flask
# --- Pour traitement séquentiel ---
import io
import contextlib

def process_all_pdfs_in_folder_with_logs_sequentiel(pdf_folder, log_func=print):
    import pandas as pd
    output_dir = os.path.join(pdf_folder, "outputs")
    os.makedirs(output_dir, exist_ok=True)

    for file in os.listdir(pdf_folder):
        if file.endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, file)
            output_excel = os.path.join(output_dir, file.replace(".pdf", "_fusion_finale_TOPDOWN_OK.xlsx"))
            log_func(f"\n📄 Traitement : {file}")
            try:
                extract_and_order_blocks(pdf_path, output_excel, log_func)
                create_fused_results_sheet(output_excel, log_func)
                create_global_summary(output_excel, log_func)
                fix_duplicate_realisations_in_excel(output_excel, log_func)
                check_duplicate_realisations_in_excel(output_excel, log_func)
            except Exception as e:
                log_func(f"❌ Erreur avec {file} : {e}")

    log_func("\n✅ Tous les fichiers PDF ont été traités.")
    log_func("\n📦 Fusion des Résumés Globaux...")

    all_data = []
    for file in os.listdir(output_dir):
        if file.endswith("_fusion_finale_TOPDOWN_OK.xlsx"):
            path = os.path.join(output_dir, file)
            try:
                df = pd.read_excel(path, sheet_name="Résumé Global")
                df["Source PDF"] = file.replace("_fusion_finale_TOPDOWN_OK.xlsx", ".pdf")
                all_data.append(df)
            except Exception as e:
                log_func(f"⚠️ Erreur avec {file} : {e}")

    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)
        final_path = os.path.join(output_dir, "Résumé_Global_Complet.xlsx")
        final_df.to_excel(final_path, index=False)
        log_func(f"\n📊 Fichier résumé final sauvegardé : {final_path}")
    else:
        log_func("❌ Aucun résumé global trouvé à fusionner.")






# --- Remplace le main par celui-ci ---
if __name__ == "__main__":
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title="Sélectionnez le dossier contenant les fichiers PDF")
    root.destroy()

    if folder_selected:
        process_all_pdfs_in_folder(folder_selected)
